"""\
Met à jour les slugs des SousCategories à partir d'un fichier NAF (Excel/CSV).

Objectif (cas d'usage):
- La base contient des SousCategories dont le slug (et parfois le nom) est basé sur un code NAF.
- On veut remplacer ces slugs par des slugs lisibles basés sur l'intitulé NAF d'un fichier
  officiel (NAF rev.2) fourni sous forme .xlsx ou .csv.

Contraintes implémentées:
- On ne traite que les codes NAF existants dans la base (Entreprise.naf_code).
- On modifie uniquement le champ `slug` des SousCategories (pas le nom), sauf si
  vous étendez la commande.
- Supporte un mode `--dry-run`.

Exemples:
    python manage.py update_naf_slugs_from_file --file data/naf.xlsx --dry-run
    python manage.py update_naf_slugs_from_file --file data/naf.xlsx
    python manage.py update_naf_slugs_from_file --file data/naf.csv --delimiter ';' --dry-run
"""

from __future__ import annotations

import csv
import pathlib
import re
from collections.abc import Iterable

from django.core.management.base import BaseCommand
from django.db import transaction
from django.db.models import Q
from django.utils.text import slugify

from foxreviews.enterprise.models import Entreprise
from foxreviews.subcategory.models import SousCategorie


_NAF_CODE_RE = re.compile(r"^\d{1,2}(?:\.\d{1,2}){0,2}[A-Z0-9]?$", re.IGNORECASE)
_NAF_FULL_RE = re.compile(r"^\d{2}\.\d{2}[A-Z0-9]$", re.IGNORECASE)
_NAF_NO_DOT_RE = re.compile(r"^\d{4}[A-Z0-9]$", re.IGNORECASE)


def _normalize_naf_code(value: str | None) -> str | None:
    if not value:
        return None

    code = str(value).strip().upper()
    if not code:
        return None

    # Excel numeric values may come as strings like "1.0" or "1.11".
    # Normalize simple trailing .0
    if re.fullmatch(r"\d+\.0", code):
        code = code[:-2]

    # Ignore headers/section markers
    if code.startswith("SECTION"):
        return None

    # Common INSEE format without dot: 6201Z
    if _NAF_NO_DOT_RE.fullmatch(code):
        return f"{code[:2]}.{code[2:]}"

    if not _NAF_CODE_RE.fullmatch(code):
        return None

    # Left-pad the first segment to 2 digits when needed:
    # "1" -> "01", "1.11" -> "01.11"
    parts = code.split(".")
    if parts and parts[0].isdigit() and len(parts[0]) == 1:
        parts[0] = f"0{parts[0]}"
        code = ".".join(parts)

    # Prefer dot format when possible
    if _NAF_FULL_RE.fullmatch(code):
        return code

    return code


def _ensure_unique_slug(base_slug: str, exclude_id) -> str:
    """Ancien helper (compat).

    ⚠️ Ne tient pas compte des collisions *entre* objets mis à jour dans le même
    `bulk_update` (car l'état DB ne change qu'à la fin). Conservez pour compat,
    mais préférez `_ensure_unique_slug_in_memory`.
    """

    slug = base_slug
    counter = 1

    while SousCategorie.objects.filter(slug=slug).exclude(id=exclude_id).exists():
        slug = f"{base_slug}-{counter}"
        counter += 1

    return slug


def _ensure_unique_slug_in_memory(*, base_slug: str, naf_code: str, reserved_slugs: set[str]) -> str:
    """Génère un slug unique en tenant compte des slugs déjà "réservés".

    Stratégie:
    - candidate = base_slug
    - si collision, candidate = f"{base_slug}-{naf_code_sans_points}"
    - si collision encore, candidate = f"{base_slug}-{naf_code_sans_points}-{i}"

    Garantit max_length=120 (champ SousCategorie.slug).
    """

    max_len = 120
    base_slug = (base_slug or "").strip("-")
    naf_suffix = re.sub(r"[^a-z0-9]+", "", (naf_code or "").lower().replace(".", ""))
    naf_suffix = naf_suffix or "naf"

    def build(base: str, suffix: str | None) -> str:
        base = (base or "").strip("-")
        if not suffix:
            return base[:max_len]

        suffix = suffix.strip("-")
        # 1 char for '-'
        cut = max_len - (len(suffix) + 1)
        if cut < 1:
            # Worst-case fallback
            return (suffix[:max_len]).strip("-")
        return f"{base[:cut].rstrip('-')}-{suffix}".strip("-")

    candidate = build(base_slug, None)
    if candidate and candidate not in reserved_slugs:
        return candidate

    candidate = build(base_slug, naf_suffix)
    if candidate and candidate not in reserved_slugs:
        return candidate

    i = 2
    while True:
        candidate = build(base_slug, f"{naf_suffix}-{i}")
        if candidate and candidate not in reserved_slugs:
            return candidate
        i += 1


def _pick_label(row: dict[str, str], prefer: str) -> str | None:
    # Try multiple possible header names (Excel exports vary)
    keys = {k.strip().lower(): k for k in row.keys() if k is not None}

    def get_by_substring(substr: str) -> str | None:
        substr = substr.lower()
        for k_norm, original in keys.items():
            if substr in k_norm:
                val = row.get(original)
                if val is None:
                    continue
                v = str(val).strip()
                if v:
                    return v
        return None

    prefer = prefer.lower().strip()

    if prefer in {"40", "40c", "40car", "40caracteres", "40 caractères"}:
        return get_by_substring("40") or get_by_substring("65") or get_by_substring("version finale")

    if prefer in {"65", "65c", "65car", "65caracteres", "65 caractères"}:
        return get_by_substring("65") or get_by_substring("40") or get_by_substring("version finale")

    # default: version finale, then 65, then 40
    return get_by_substring("version finale") or get_by_substring("65") or get_by_substring("40")


def _iter_rows_from_xlsx(path: pathlib.Path, sheet: str | None) -> Iterable[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    header_names = [str(h).strip() if h is not None else "" for h in headers]

    def to_dict(values) -> dict[str, str]:
        out: dict[str, str] = {}
        for idx, name in enumerate(header_names):
            if not name:
                continue
            cell = values[idx] if idx < len(values) else None
            if cell is None:
                continue
            out[name] = str(cell)
        return out

    for values in rows:
        if not values:
            continue
        row = to_dict(values)
        if row:
            yield row


def _iter_rows_from_xls(path: pathlib.Path, sheet: str | None) -> Iterable[dict[str, str]]:
    """Lit un fichier Excel .xls (BIFF) via xlrd (optionnel)."""
    try:
        import xlrd  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "Lecture .xls indisponible. Installez 'xlrd' (pip install xlrd) "
            "ou convertissez le fichier en .xlsx."
        ) from exc

    book = xlrd.open_workbook(str(path))
    ws = book.sheet_by_name(sheet) if sheet else book.sheet_by_index(0)

    if ws.nrows == 0:
        return []

    header_names = [str(h).strip() if h is not None else "" for h in ws.row_values(0)]

    for rowx in range(1, ws.nrows):
        values = ws.row_values(rowx)
        out: dict[str, str] = {}
        for idx, name in enumerate(header_names):
            if not name:
                continue
            if idx >= len(values):
                continue
            cell = values[idx]
            if cell is None or cell == "":
                continue
            out[name] = str(cell)
        if out:
            yield out


def _iter_rows_from_csv(path: pathlib.Path, delimiter: str) -> Iterable[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            if not row:
                continue
            yield {k: (v if v is not None else "") for k, v in row.items() if k is not None}


def _get_code_from_row(row: dict[str, str]) -> str | None:
    keys = {k.strip().lower(): k for k in row.keys() if k is not None}

    # Typical header "Code" or "code"
    for candidate in ("code", "naf", "code naf"):
        if candidate in keys:
            v = str(row.get(keys[candidate], "")).strip()
            if v:
                return v

    # Fallback: try any column that contains 'code'
    for k_norm, original in keys.items():
        if "code" in k_norm:
            v = str(row.get(original, "")).strip()
            if v:
                return v

    return None


def _resolve_sous_categorie_for_naf(naf_code: str) -> SousCategorie | None:
    """Heuristiques pour retrouver la SousCategorie liée à un code NAF."""

    naf_no_dot = naf_code.replace(".", "")

    slug_candidates = {
        slugify(naf_code),
        slugify(naf_no_dot),
        slugify(f"Activité {naf_code}"),
        slugify(f"Activite {naf_code}"),
        slugify(f"Activité {naf_no_dot}"),
        slugify(f"Activite {naf_no_dot}"),
    }
    name_candidates = {
        naf_code,
        naf_no_dot,
        f"Activité {naf_code}",
        f"Activite {naf_code}",
        f"Activité {naf_no_dot}",
        f"Activite {naf_no_dot}",
    }

    qs = SousCategorie.objects.filter(
        Q(slug__in=[s for s in slug_candidates if s])
        | Q(nom__in=[n for n in name_candidates if n])
        | Q(description__startswith=f"NAF {naf_code}")
        | Q(description__startswith=f"NAF {naf_no_dot}")
        | Q(description__icontains=f"NAF {naf_code}")
        | Q(description__icontains=f"NAF {naf_no_dot}")
    )

    candidates = list(qs[:10])
    if not candidates:
        return None

    if len(candidates) == 1:
        return candidates[0]

    # Prefer the very explicit "Activité <code>" naming
    for c in candidates:
        if c.nom.strip() in {f"Activité {naf_code}", f"Activite {naf_code}"}:
            return c
    for c in candidates:
        if c.nom.strip() in {naf_code, naf_no_dot}:
            return c

    # Otherwise, keep deterministic: smallest id (stable-ish)
    return sorted(candidates, key=lambda sc: str(sc.id))[0]


class Command(BaseCommand):
    help = "Met à jour les slugs des sous-catégories depuis un fichier NAF (Excel/CSV)"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            required=True,
            help="Chemin vers le fichier .xlsx ou .csv (NAF rev.2)",
        )
        parser.add_argument(
            "--sheet",
            default=None,
            help="Nom de l'onglet Excel (si .xlsx). Par défaut: onglet actif.",
        )
        parser.add_argument(
            "--delimiter",
            default=",",
            help="Délimiteur CSV (par défaut ','). Exemple: ';'",
        )
        parser.add_argument(
            "--prefer",
            default="40",
            help=(
                "Colonne d'intitulé à préférer pour construire le slug: '40', '65' ou 'final'. "
                "Par défaut: '40' (plus court)."
            ),
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simulation: affiche ce qui serait modifié sans écrire en base.",
        )
        parser.add_argument(
            "--limit",
            type=int,
            default=None,
            help="Limiter le nombre de mises à jour (debug).",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = pathlib.Path(options["file"]).expanduser().resolve()
        sheet = options.get("sheet")
        delimiter = options.get("delimiter") or ","
        prefer = options.get("prefer") or "40"
        dry_run = bool(options.get("dry_run"))
        limit = options.get("limit")

        if not file_path.exists():
            self.stdout.write(self.style.ERROR(f"❌ Fichier introuvable: {file_path}"))
            return

        if file_path.suffix.lower() not in {".xlsx", ".xls", ".csv"}:
            self.stdout.write(self.style.ERROR("❌ Formats supportés: .xlsx, .xls, .csv"))
            return

        # 1) Récupérer les codes NAF présents dans la base
        self.stdout.write("🔍 Chargement des codes NAF existants en base...")
        naf_in_db_raw = Entreprise.objects.values_list("naf_code", flat=True).distinct()

        naf_in_db: set[str] = set()
        for code in naf_in_db_raw.iterator(chunk_size=5000):
            norm = _normalize_naf_code(code)
            if norm:
                naf_in_db.add(norm)

        self.stdout.write(f"   → {len(naf_in_db)} codes NAF uniques détectés")

        # 2) Lire le fichier et construire une table code -> label
        self.stdout.write(f"📄 Lecture du fichier: {file_path.name}")

        try:
            if file_path.suffix.lower() == ".xlsx":
                rows = _iter_rows_from_xlsx(file_path, sheet)
            elif file_path.suffix.lower() == ".xls":
                rows = _iter_rows_from_xls(file_path, sheet)
            else:
                rows = _iter_rows_from_csv(file_path, delimiter)
        except RuntimeError as exc:
            self.stdout.write(self.style.ERROR(f"❌ {exc}"))
            return

        code_to_label: dict[str, str] = {}
        total_rows = 0
        ignored_rows = 0

        for row in rows:
            total_rows += 1
            raw_code = _get_code_from_row(row)
            norm_code = _normalize_naf_code(raw_code)
            if not norm_code:
                ignored_rows += 1
                continue

            # Ne garder que les codes NAF existants dans la base
            if norm_code not in naf_in_db:
                continue

            label = _pick_label(row, prefer)
            if not label:
                ignored_rows += 1
                continue

            code_to_label[norm_code] = label

        self.stdout.write(
            f"   → {total_rows} lignes lues, {ignored_rows} ignorées, {len(code_to_label)} codes utiles (présents en base)"
        )

        if not code_to_label:
            self.stdout.write(self.style.WARNING("⚠️  Aucun code NAF du fichier ne matche la base."))
            return

        # 3) Préparer les updates
        to_update: list[SousCategorie] = []
        updated = 0
        unchanged = 0
        missing_subcategory = 0

        # Important: éviter les collisions entre plusieurs updates dans un même bulk_update.
        # On maintient un set des slugs "réservés" correspondant à l'état final.
        reserved_slugs: set[str] = set(SousCategorie.objects.values_list("slug", flat=True))

        for naf_code, label in code_to_label.items():
            sous_cat = _resolve_sous_categorie_for_naf(naf_code)
            if not sous_cat:
                missing_subcategory += 1
                continue

            base_slug = slugify(label)
            if not base_slug:
                continue

            # Le slug courant peut être "libéré" si on le modifie.
            reserved_slugs.discard(sous_cat.slug)

            new_slug = _ensure_unique_slug_in_memory(
                base_slug=base_slug,
                naf_code=naf_code,
                reserved_slugs=reserved_slugs,
            )

            if sous_cat.slug == new_slug:
                # Pas de changement, on re-réserve le slug courant.
                reserved_slugs.add(sous_cat.slug)
                unchanged += 1
                continue

            old_slug = sous_cat.slug
            sous_cat.slug = new_slug
            to_update.append(sous_cat)
            updated += 1

            # Réserver le nouveau slug pour éviter les collisions avec les prochains.
            reserved_slugs.add(new_slug)

            if limit is not None and updated >= limit:
                break

            if dry_run:
                self.stdout.write(f"   DRY-RUN: {naf_code:10} → {old_slug} -> {new_slug}")

        # 4) Appliquer
        if dry_run:
            self.stdout.write(self.style.WARNING("\n⚠️  DRY-RUN: aucune modification écrite en base"))
        else:
            if to_update:
                self.stdout.write(f"\n📝 Application des mises à jour: {len(to_update)} slugs...")
                SousCategorie.objects.bulk_update(to_update, ["slug"], batch_size=200)
                self.stdout.write(self.style.SUCCESS("✅ Slugs mis à jour"))

        # 5) Résumé
        self.stdout.write("\n" + "=" * 80)
        self.stdout.write("📊 RÉSUMÉ update_naf_slugs_from_file")
        self.stdout.write("=" * 80)
        self.stdout.write(f"Codes NAF retenus (fichier ∩ base): {len(code_to_label)}")
        self.stdout.write(f"SousCategories modifiées: {updated}")
        self.stdout.write(f"SousCategories inchangées: {unchanged}")
        self.stdout.write(f"SousCategories introuvables pour un NAF: {missing_subcategory}")
        self.stdout.write("=" * 80)

        if not dry_run and updated > 0:
            self.stdout.write(
                self.style.WARNING(
                    "\n⚠️  Attention: si vous avez un mapping NAF→slug (ex: naf_mapping.py), "
                    "il devra être mis à jour pour refléter les nouveaux slugs."
                )
            )
